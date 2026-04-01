import csv
from pathlib import Path
import xlrd
from openpyxl import load_workbook
from pypdf import PdfReader


class DocumentReader:
    def read_knowledge_file(self, file_path: Path) -> str:
        suffix = file_path.suffix.lower()
        if suffix in ['.md', '.txt']:
            return file_path.read_text(encoding='utf-8')
        if suffix == '.pdf':
            return self._read_pdf(file_path)
        if suffix in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            return self._read_excel(file_path)
        if suffix == '.xls':
            return self._read_xls(file_path)
        if suffix == '.csv':
            return self._read_csv(file_path)
        return ''

    def _read_pdf(self, file_path: Path) -> str:
        reader = PdfReader(str(file_path))
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or '')
        return '\n'.join(pages)

    def _read_excel(self, file_path: Path) -> str:
        workbook = load_workbook(filename=str(file_path), data_only=True, read_only=True)
        rows_text = []
        for sheet in workbook.worksheets:
            sheet_rows = []
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip() != '']
                if values:
                    sheet_rows.append(' | '.join(values))
            if sheet_rows:
                rows_text.append(f"# Sheet: {sheet.title}\n" + '\n'.join(sheet_rows))
        workbook.close()
        return '\n\n'.join(rows_text)

    def _read_xls(self, file_path: Path) -> str:
        workbook = xlrd.open_workbook(str(file_path))
        rows_text = []
        for sheet in workbook.sheets():
            sheet_rows = []
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                values = [str(cell).strip() for cell in row if str(cell).strip() != '']
                if values:
                    sheet_rows.append(' | '.join(values))
            if sheet_rows:
                rows_text.append(f"# Sheet: {sheet.name}\n" + '\n'.join(sheet_rows))
        return '\n\n'.join(rows_text)

    def _read_csv(self, file_path: Path) -> str:
        rows = []
        with file_path.open('r', encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            for row in reader:
                cleaned = [cell.strip() for cell in row if cell is not None and cell.strip() != '']
                if cleaned:
                    rows.append(' | '.join(cleaned))
        return '\n'.join(rows)